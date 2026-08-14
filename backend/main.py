import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
import io
import argparse

from fastapi import FastAPI, Response, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware

from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.metrics import silhouette_score, davies_bouldin_score
import uvicorn
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================================================
# FASTAPI
# =========================================================
app = FastAPI(
    title="API Evaluasi Efektivitas Delegasi Tugas"
)

# =========================================================
# CORS
# =========================================================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# ROOT
# =========================================================
@app.get("/", include_in_schema=False)
async def root():
    return RedirectResponse(url="/docs")


# =========================================================
# REQUIRED DATASET COLUMNS
# =========================================================
REQUIRED_COLUMNS = [
    'sprint_id',
    'role',
    'assignee',
    'story_point',
    'complexity_score',
    'risk_score',
    'dependency_score',
    'uncertainty_score',
    'volume_score',
    'task_duration_hours',
    'reopen_count',
    'role_capacity'
]


# =========================================================
# PREPROCESSING + KMEANS
# =========================================================
def perform_clustering(df, k=3):

    # =====================================================
    # PHASE 1: VALIDASI DATASET & PREPROCESSING
    # =====================================================
    missing_cols = [
        col for col in REQUIRED_COLUMNS
        if col not in df.columns
    ]

    if missing_cols:
        raise ValueError(
            f"Kolom tidak ditemukan: {', '.join(missing_cols)}"
        )

    if df.empty:
        raise ValueError("Dataset kosong.")

    # CLEANING
    df['assignee'] = df['assignee'].astype(str).str.strip()
    df['story_point'] = pd.to_numeric(df['story_point'], errors='coerce')
    df = df.dropna(subset=['assignee', 'story_point'])

    # =====================================================
    # PHASE 2 & 3: AGREGASI PER KARYAWAN & FEATURE
    # =====================================================
    df_user = df.groupby('assignee').agg({
        'story_point': 'sum',
        'role': 'first' # Data tambahan/context, tidak diikutkan clustering
    }).reset_index()

    df_user.rename(columns={'story_point': 'total_history_point'}, inplace=True)
    
    # Feature matriks hanya history point per skripsi
    X = df_user[['total_history_point']]

    # =====================================================
    # PHASE 4: K-MEANS K=3
    # =====================================================
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Harus K=3 secara metodologis
    kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
    df_user['cluster'] = kmeans.fit_predict(X_scaled)

    # =====================================================
    # PHASE 5: CENTROID & INTERPRETASI CLUSTER
    # =====================================================
    centroids = kmeans.cluster_centers_
    original_centroids = scaler.inverse_transform(centroids).flatten()
    
    # Mengurutkan index centroid dari terendah ke tertinggi
    sorted_idx = np.argsort(original_centroids)
    
    cluster_mapping = {
        sorted_idx[0]: "Rendah",
        sorted_idx[1]: "Sedang",
        sorted_idx[2]: "Tinggi"
    }

    df_user['workload_category'] = df_user['cluster'].map(cluster_mapping)

    cluster_summary = df_user.groupby('cluster').agg(
        total_history_point_mean=('total_history_point', 'mean'),
        employee_count=('assignee', 'count')
    ).reset_index()
    
    cluster_summary['workload_category'] = cluster_summary['cluster'].map(cluster_mapping)

    # =====================================================
    # PHASE 6: DAVIES-BOULDIN INDEX
    # =====================================================
    # Skripsi menggunakan DBI, semakin kecil semakin baik.
    dbi = davies_bouldin_score(X_scaled, df_user['cluster'])

    return (
        df_user,
        cluster_summary,
        dbi,
        X_scaled
    )


# =========================================================
# ELBOW METHOD
# =========================================================
def calculate_elbow(X_scaled):

    inertia = []

    K = range(1, 11)

    for k in K:

        model = KMeans(
            n_clusters=k,
            random_state=42,
            n_init=10
        )

        model.fit(X_scaled)

        inertia.append(model.inertia_)

    return list(K), inertia


# =========================================================
# GENERATE VISUALIZATION
# =========================================================
def generate_preview(
    df_result,
    dbi_value,
    output_path='results/latest_preview.png'
):

    sns.set_theme(style="whitegrid")

    plt.figure(figsize=(16, 8))

    # =====================================================
    # SCATTERPLOT
    # =====================================================
    sns.scatterplot(
        data=df_result,
        x='assignee',
        y='total_history_point',
        hue='workload_category',
        palette={'Rendah': '#10b981', 'Sedang': '#f59e0b', 'Tinggi': '#ef4444'},
        s=140,
        alpha=0.85,
        edgecolor='black'
    )

    plt.title(
        'Pengelompokan Beban Kerja Karyawan',
        fontsize=18,
        fontweight='bold'
    )

    plt.xlabel('Assignee')
    plt.ylabel('Total History Point')

    plt.xticks(rotation=45)

    # =====================================================
    # SUMMARY
    # =====================================================
    total_tasks = len(df_result)

    overload_tasks = len(
        df_result[df_result['workload_category'] == 'Tinggi']
    )

    high_rework = 0

    summary_text = (
        f"📊 Delegation Analytics\n\n"
        f"Total Task: {total_tasks}\n"
        f"Davies-Bouldin Index: {dbi_value:.4f}\n"
        f"Karyawan Beban Tinggi: {overload_tasks}\n\n"
        f"💡 Insight:\n"
        f"Klaster digunakan untuk\n"
        f"evaluasi pemerataan beban\n"
        f"kerja secara objektif\n"
        f"menggunakan K-Means."
    )

    plt.text(
        1.02,
        0.25,
        summary_text,
        transform=plt.gca().transAxes,
        fontsize=11,
        bbox=dict(
            boxstyle='round',
            facecolor='white',
            edgecolor='gray'
        )
    )

    plt.tight_layout(rect=[0, 0, 0.82, 1])

    os.makedirs(
        os.path.dirname(output_path),
        exist_ok=True
    )

    plt.savefig(
        output_path,
        dpi=300
    )

    plt.close()


# =========================================================
# TEMPLATE CSV
# =========================================================
@app.get("/template")
async def get_template():

    template_data = {
        'sprint_id': ['Sprint-1', 'Sprint-1'],
        'role': ['Backend', 'Frontend'],
        'assignee': ['Rahman', 'Siti'],
        'story_point': [8, 3],
        'complexity_score': [5, 2],
        'risk_score': [4, 2],
        'dependency_score': [3, 1],
        'uncertainty_score': [4, 1],
        'volume_score': [5, 2],
        'task_duration_hours': [48, 12],
        'reopen_count': [2, 0],
        'role_capacity': [35, 30]
    }

    df_template = pd.DataFrame(
        template_data
    )

    stream = io.StringIO()

    df_template.to_csv(
        stream,
        index=False
    )

    return Response(
        content=stream.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=template_evaluasi_delegasi.csv"
        }
    )


# =========================================================
# DRY RUN / VALIDASI
# =========================================================
@app.post("/dry-run")
async def dry_run(file: UploadFile = File(...)):
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=400,
            detail="Hanya file CSV diperbolehkan."
        )
        
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        
        missing_cols = [
            col for col in REQUIRED_COLUMNS
            if col not in df.columns
        ]

        if missing_cols:
            raise ValueError(
                f"Kolom tidak ditemukan: {', '.join(missing_cols)}"
            )

        if df.empty:
            raise ValueError("Dataset kosong.")
            
        numeric_cols = [
            'story_point', 'complexity_score', 'risk_score',
            'dependency_score', 'uncertainty_score', 'volume_score',
            'task_duration_hours', 'reopen_count', 'role_capacity'
        ]

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        if df[numeric_cols].isnull().any().any():
            raise ValueError("Terdapat data numerik invalid / kosong.")
            
        return {
            "status": "success", 
            "message": "File valid. Semua kolom yang dibutuhkan tersedia dan format data sesuai."
        }
        
    except ValueError as ve:
        raise HTTPException(
            status_code=400,
            detail=str(ve)
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Server Error: {str(e)}"
        )



# =========================================================
# UPLOAD CSV
# =========================================================
@app.post("/upload")
async def upload_file(
    request: Request,
    file: UploadFile = File(...)
):

    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=400,
            detail="Hanya file CSV diperbolehkan."
        )

    try:

        contents = await file.read()

        df = pd.read_csv(
            io.BytesIO(contents)
        )

        (
            df_result,
            cluster_summary,
            dbi_val,
            X_scaled
        ) = perform_clustering(df)

        # =================================================
        # GENERATE PREVIEW
        # =================================================
        preview_path = 'results/latest_preview.png'

        generate_preview(
            df_result,
            dbi_val,
            preview_path
        )

        # =================================================
        # ELBOW METHOD
        # =================================================
        K, inertia = calculate_elbow(
            X_scaled
        )

        # =================================================
        # ROLE ANALYSIS
        # =================================================
        role_analysis = (
            df_result
            .groupby(['role', 'cluster'])
            .size()
            .reset_index(name='total_task')
        )

        # =================================================
        # ASSIGNEE ANALYSIS
        # =================================================
        assignee_analysis = (
            df_result
            .groupby(['assignee', 'cluster'])
            .size()
            .reset_index(name='total_task')
        )

        # =================================================
        # ROLE WORKLOAD
        # =================================================
        role_workload = (
            df_result
            .groupby('role')['total_history_point']
            .sum()
            .reset_index(name='total_story_point')
        )

        # =================================================
        # ASSIGNEE WORKLOAD
        # =================================================
        assignee_workload = (
            df_result[['assignee', 'total_history_point', 'workload_category']]
        )

        # =================================================
        # EXCEL EXPORT
        # =================================================
        excel_path = 'results/hasil_evaluasi_delegasi.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            cluster_summary.to_excel(writer, sheet_name='Ringkasan Klaster', startrow=5, index=False)
            df_result.to_excel(writer, sheet_name='Detail Karyawan', index=False)

            workbook = writer.book
            
            # Format Ringkasan Klaster
            ws1 = writer.sheets['Ringkasan Klaster']
            
            thin_border = Border(
                left=Side(style='thin', color="D1D5DB"), 
                right=Side(style='thin', color="D1D5DB"), 
                top=Side(style='thin', color="D1D5DB"), 
                bottom=Side(style='thin', color="D1D5DB")
            )
            
            # Add Stats Cards
            title_font = Font(bold=True, size=14, color="1E3A8A")
            label_font = Font(bold=True, color="6B7280")
            value_font = Font(bold=True, size=16, color="4F46E5")
            
            ws1['A1'] = "HASIL ANALISIS BEBAN KERJA"
            ws1['A1'].font = title_font
            
            ws1['A3'] = "USER TERANALISIS"
            ws1['A3'].font = label_font
            ws1['A4'] = len(df_result)
            ws1['A4'].font = value_font
            ws1['A4'].alignment = Alignment(horizontal='left')
            
            ws1['B3'] = "DAVIES-BOULDIN INDEX"
            ws1['B3'].font = label_font
            ws1['B4'] = round(dbi_val, 4)
            ws1['B4'].font = value_font
            ws1['B4'].alignment = Alignment(horizontal='left')
            
            ws1['C3'] = "JUMLAH CLUSTER"
            ws1['C3'].font = label_font
            ws1['C4'] = len(cluster_summary)
            ws1['C4'].font = value_font
            ws1['C4'].alignment = Alignment(horizontal='left')
            
            header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws1["6:6"]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            for r in range(6, 6 + len(cluster_summary) + 1):
                for c in range(1, len(cluster_summary.columns) + 1):
                    ws1.cell(row=r, column=c).border = thin_border
            
            # Adjust column width for Ringkasan Klaster
            for col in ws1.columns:
                max_length = 0
                column = col[0].column_letter 
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws1.column_dimensions[column].width = max_length + 2
            
            # Insert Chart to Ringkasan Klaster
            if os.path.exists(preview_path):
                img = Image(preview_path)
                # Resize image slightly to fit nicely
                img.width = 650
                img.height = 450
                # Insert at cell A14
                ws1.add_image(img, 'A14')
            
            # Add Recommendations
            ws1['A39'] = "💡 Rekomendasi Pemerataan Beban Kerja:"
            ws1['A39'].font = Font(bold=True, size=14, color="1E3A8A")
            
            row = 41
            for index, row_data in cluster_summary.iterrows():
                cat = row_data['workload_category']
                cluster_id = row_data['cluster']
                
                users = df_result[df_result['cluster'] == cluster_id]['assignee'].tolist()
                users_str = ", ".join(users)
                
                start_col, end_col = 1, 7
                
                if cat == 'Tinggi':
                    bg_color = "FEF2F2"
                    border_color = "FCA5A5"
                    ws1[f'A{row}'] = f"⚠️ Karyawan Beban Tinggi:"
                    ws1[f'A{row}'].font = Font(bold=True, color="EF4444", size=12)
                    ws1[f'A{row+1}'] = f"User: {users_str}"
                    ws1[f'A{row+1}'].font = Font(bold=True)
                    ws1[f'A{row+2}'] = "Peringatan: Karyawan ini memikul beban tugas tertinggi. Diperlukan pemerataan tugas (delegasi ulang) untuk menghindari bottleneck."
                    ws1[f'A{row+2}'].font = Font(color="7F1D1D")
                elif cat == 'Sedang':
                    bg_color = "FFFBEB"
                    border_color = "FCD34D"
                    ws1[f'A{row}'] = f"⚖️ Karyawan Beban Sedang:"
                    ws1[f'A{row}'].font = Font(bold=True, color="F59E0B", size=12)
                    ws1[f'A{row+1}'] = f"User: {users_str}"
                    ws1[f'A{row+1}'].font = Font(bold=True)
                    ws1[f'A{row+2}'] = "Beban kerja karyawan ini relatif seimbang dan proporsional dengan kapasitas tim."
                    ws1[f'A{row+2}'].font = Font(color="78350F")
                elif cat == 'Rendah':
                    bg_color = "ECFDF5"
                    border_color = "6EE7B7"
                    ws1[f'A{row}'] = f"🚀 Karyawan Beban Rendah:"
                    ws1[f'A{row}'].font = Font(bold=True, color="10B981", size=12)
                    ws1[f'A{row+1}'] = f"User: {users_str}"
                    ws1[f'A{row+1}'].font = Font(bold=True)
                    ws1[f'A{row+2}'] = "Karyawan ini masih memiliki kapasitas beban yang minim. Sangat direkomendasikan untuk mendelegasikan tugas-tugas tambahan kepada karyawan ini."
                    ws1[f'A{row+2}'].font = Font(color="064E3B")
                
                fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                
                for r_idx in range(row, row+3):
                    for c_idx in range(start_col, end_col + 1):
                        cell = ws1.cell(row=r_idx, column=c_idx)
                        cell.fill = fill
                        
                        top_side = Side(style='medium', color=border_color) if r_idx == row else None
                        bottom_side = Side(style='medium', color=border_color) if r_idx == row+2 else None
                        left_side = Side(style='medium', color=border_color) if c_idx == start_col else None
                        right_side = Side(style='medium', color=border_color) if c_idx == end_col else None
                        
                        cell.border = Border(top=top_side, bottom=bottom_side, left=left_side, right=right_side)
                
                row += 4
                
            # Format Detail Karyawan
            ws2 = writer.sheets['Detail Karyawan']
            header_fill2 = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            for cell in ws2["1:1"]:
                cell.fill = header_fill2
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            for r in range(1, 1 + len(df_result) + 1):
                for c in range(1, len(df_result.columns) + 1):
                    ws2.cell(row=r, column=c).border = thin_border
            
            # Adjust column width for Detail Karyawan
            for col in ws2.columns:
                max_length = 0
                column = col[0].column_letter 
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws2.column_dimensions[column].width = max_length + 2

        # =================================================
        # URL
        # =================================================
        base_url = str(request.base_url)

        preview_url = f"{base_url}preview"

        download_url = f"{base_url}download"

        excel_url = f"{base_url}download-excel"

        return {

            "status": "success",

            "message":
                "Clustering beban kerja berhasil dilakukan.",

            "davies_bouldin_index":
                round(dbi_val, 4),

            "cluster_summary":
                cluster_summary
                .reset_index()
                .to_dict(orient='records'),

            "role_analysis":
                role_analysis
                .to_dict(orient='records'),

            "assignee_analysis":
                assignee_analysis
                .to_dict(orient='records'),

            "role_workload":
                role_workload
                .to_dict(orient='records'),

            "assignee_workload":
                assignee_workload
                .to_dict(orient='records'),

            "elbow_method": {
                "k_values": K,
                "inertia": inertia
            },

            "preview_url": preview_url,

            "excel_url": excel_url,

            "data":
                df_result.to_dict(
                    orient='records'
                )
        }

    except ValueError as ve:

        raise HTTPException(
            status_code=400,
            detail=str(ve)
        )

    except Exception as e:

        raise HTTPException(
            status_code=500,
            detail=f"Server Error: {str(e)}"
        )


# =========================================================
# PREVIEW IMAGE
# =========================================================
@app.get("/preview")
async def get_preview():

    preview_path = 'results/latest_preview.png'

    if not os.path.exists(preview_path):

        raise HTTPException(
            status_code=404,
            detail="Preview belum tersedia."
        )

    return FileResponse(preview_path)


# =========================================================
# DOWNLOAD EXCEL
# =========================================================
@app.get("/download-excel")
async def download_excel():

    excel_path = 'results/hasil_evaluasi_delegasi.xlsx'

    if not os.path.exists(excel_path):
        raise HTTPException(
            status_code=404,
            detail="File Excel tidak ditemukan."
        )

    return FileResponse(
        path=excel_path,
        filename="hasil_evaluasi_delegasi.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================================
# RUN LOCAL
# =========================================================
def run_local():

    print("=== Evaluasi Efektivitas Delegasi ===")

    file_path = 'kmeans_dataset_updated.csv'

    if not os.path.exists(file_path):

        print(f"File tidak ditemukan: {file_path}")

        return

    df = pd.read_csv(file_path)

    (
        df_result,
        cluster_summary,
        dbi_val,
        X_scaled
    ) = perform_clustering(df)

    # =====================================================
    # PRINT ANALYSIS
    # =====================================================
    print("\n=== DAVIES-BOULDIN INDEX ===")

    print(round(dbi_val, 4))

    print("\n=== CLUSTER SUMMARY ===")

    print(cluster_summary)

    # =====================================================
    # VISUALISASI
    # =====================================================
    plt.figure(figsize=(14, 7))

    sns.scatterplot(
        data=df_result,
        x='assignee',
        y='total_history_point',
        hue='workload_category',
        palette={'Rendah': '#10b981', 'Sedang': '#f59e0b', 'Tinggi': '#ef4444'},
        s=120
    )

    plt.title(
        'Pengelompokan Beban Kerja Karyawan'
    )

    plt.xlabel('Assignee')

    plt.ylabel('Total History Point')

    plt.xticks(rotation=45)

    # plt.show() # Disable for headless execution

    # =====================================================
    # SAVE RESULT
    # =====================================================
    os.makedirs(
        'results',
        exist_ok=True
    )

    output_path = (
        'results/hasil_evaluasi_delegasi.csv'
    )

    df_result.to_csv(
        output_path,
        index=False
    )

    print(f"\nHasil disimpan ke: {output_path}")


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--api",
        action="store_true",
        help="Jalankan sebagai API"
    )

    args = parser.parse_args()

    if args.api:

        print("API Running...")
        print("Swagger Docs: http://127.0.0.1:8000/docs")

        uvicorn.run(
            app,
            host="127.0.0.1",
            port=8000
        )

    else:

        run_local()